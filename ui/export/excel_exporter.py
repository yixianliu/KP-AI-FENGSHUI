"""
Excel 导出器
支持按"可选章节"导出：基本信息 / 命局类型 / 四柱八字 / 五行分析 /
十神分析 / 大运流年（含起运）/ 运程总结（事业/财运/健康/感情）/
吉凶批注 / AI 智能分析。调用方已用 filter_export_data
过滤 data，本导出器再按数据键是否存在逐项渲染。
"""
from typing import Dict, Any
from .base_exporter import BaseExporter, has_chapter

# 惰性导入：未安装 openpyxl 时仅置为 None，不影响本模块导入与 app 启动；
# 只有在真正实例化 / 调用 Excel 导出时才提示安装。
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    openpyxl = None
    Font = Alignment = PatternFill = Border = Side = None

# AI 字段 -> 中文标题（与 PDF 导出保持一致）
_AI_SECTIONS = [
    # 与 core.analysis_storage._JSON_SCHEMAS 对齐（三种类型的并集，仅保留列表型字段；
    # final_verdict / disclaimer 为段落型，由各自面板/导出分支单独处理）。
    ('personality', '性格特质'),
    ('career', '事业财运'),
    ('relationships', '婚姻感情'),
    ('health', '健康注意'),
    ('four_pillars_detail', '四柱详细解读'),
    ('analysis', '卦象/课体分析'),
    ('hexagram_interpretations', '卦爻解释'),
    ('timing', '应期时机'),
    ('scenario_advice', '场景化建议'),
    ('advice', '行动建议'),
    ('historical_cases', '历史案例'),
    ('probability_stats', '概率统计'),
]


class ExcelExporter(BaseExporter):
    """Excel 导出器"""

    def __init__(self):
        """初始化 Excel 导出器，预先构建整个工作表复用的样式对象集合。

        openpyxl 的 Font/Alignment/Border/PatternFill 都是不可变的样式对象，
        可以被多个单元格共享引用，所以在构造期建好一份挂到实例上，
        避免每写一个单元格就 new 一组样式（数千单元格时开销明显）。

        Raises:
            RuntimeError: 未安装 openpyxl 依赖时抛出，提示安装命令
        """
        if openpyxl is None:
            raise RuntimeError(
                "缺少依赖 openpyxl，无法导出 Excel，请先安装：pip install openpyxl"
            )
        # 字体统一用微软雅黑，保证中文在 Windows 上不会退化成默认西文字体；
        # 颜色值为 openpyxl 要求的 aRGB 八位十六进制（前两位是 alpha）
        self.styles = {
            'title': Font(name='微软雅黑', size=14, bold=True, color='FF333333'),
            'header': Font(name='微软雅黑', size=12, bold=True, color='FF5D4037'),
            'content': Font(name='微软雅黑', size=11, color='FF333333'),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='FFD4AF37'),
                right=Side(style='thin', color='FFD4AF37'),
                top=Side(style='thin', color='FFD4AF37'),
                bottom=Side(style='thin', color='FFD4AF37')
            ),
            'fill_gold': PatternFill(start_color='FFD4AF37', end_color='FFD4AF37', fill_type='solid'),
            'fill_light': PatternFill(start_color='FFFFF8E7', end_color='FFFFF8E7', fill_type='solid')
        }

    def export(self, data: Dict[str, Any], file_path: str) -> bool:
        """把排盘结果写入单个工作表并保存为 .xlsx。

        实现 BaseExporter.export 契约。所有章节纵向堆叠在同一张 sheet 上，
        用 row 游标在各 _add_* 方法之间传递当前写入位置：有返回值的方法
        直接返回下一空行行号，无返回值的方法则由调用处按数据条数手工累加，
        两种方式都要额外留出空行做章节间隔。

        Args:
            data: 排盘结果字典，可能包含 basic_info / bazi_types / bazi /
                wuxing / shishen / dayun / liunian / yuncheng / mingli /
                analysis / ai_analysis / meihua_data / meihua_ai /
                liuren_data / liuren_ai / zonghe 等键
            file_path: 目标 xlsx 文件路径

        Returns:
            bool，保存成功返回 True；写入过程中出现异常时打印错误并返回 False

        Raises:
            RuntimeError: 未安装 openpyxl 时抛出（此分支不被 try 覆盖，
                属于环境问题而非导出失败，需让调用方感知）
        """
        if openpyxl is None:
            raise RuntimeError(
                "缺少依赖 openpyxl，无法导出 Excel，请先安装：pip install openpyxl"
            )
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '八字排盘结果'

            # row 为全局写入游标，从 1 开始（openpyxl 行列号均以 1 为起点）
            row = 1
            ws.merge_cells(f'A{row}:B{row}')
            title_cell = ws.cell(row=row, column=1, value='八字排盘分析结果')
            title_cell.font = self.styles['title']
            title_cell.alignment = self.styles['center']
            # +2 而非 +1：跳过一整行留白，让主标题与首个章节拉开距离
            row += 2

            # 基本信息
            if has_chapter(data, 'basic_info'):
                row = self._add_section(ws, row, '基本信息', data.get('basic_info', {})) + 2

            # 命局类型（含用神喜忌）
            if has_chapter(data, 'bazi_types'):
                row = self._add_bazi_types_section(ws, row, data.get('bazi_types', {})) + 2

            # 四柱八字
            if has_chapter(data, 'bazi'):
                row = self._add_section(ws, row, '四柱八字', data.get('bazi', {})) + 2

            # 五行分析
            if has_chapter(data, 'wuxing'):
                row = self._add_wuxing_section(ws, row, '五行分析', data.get('wuxing', {})) + 2

            # 十神分析
            if has_chapter(data, 'shishen'):
                row = self._add_shishen_section(ws, row, '十神分析', data.get('shishen', {})) + 2

            # 大运流年
            if has_chapter(data, 'yunshi'):
                row = self._add_yunshi_section(ws, row, data) + 2

            # 运程总结（事业 / 财运 / 健康 / 感情）
            if has_chapter(data, 'yuncheng'):
                row = self._add_yuncheng_section(ws, row, data.get('yuncheng', {})) + 2

            # 神煞
            # 神煞属于 filter_export_data 中始终保留的旧字段，不在 CHAPTERS
            # 章节清单里，因此这里不走 has_chapter 而是直接判断列表是否为空
            mingli = data.get('mingli', {}) or {}
            shensha = mingli.get('shensha', [])
            if shensha:
                self._add_shensha_section(ws, row, '神煞', shensha)
                # 该方法无返回值，游标只能按「标题 1 行 + 数据 N 行 + 间隔」自行推算
                row += len(shensha) + 3

            # 吉凶批注
            analysis = data.get('analysis', []) or []
            if analysis:
                self._add_analysis_section(ws, row, '吉凶批注', analysis)
                row += len(analysis) + 3

            # AI 智能分析
            if has_chapter(data, 'ai_analysis'):
                self._add_ai_section(ws, row, data.get('ai_analysis', {}))

            # 梅花易数
            if has_chapter(data, 'meihua'):
                self._add_meihua_section(ws, row, data.get('meihua_data', {}) or {},
                                          data.get('meihua_ai', {}) or {})

            # 大六壬
            if has_chapter(data, 'liuren'):
                self._add_liuren_section(ws, row, data.get('liuren_data', {}) or {},
                                          data.get('liuren_ai', {}) or {})

            # 综合建议（融合三方结论）
            if has_chapter(data, 'zonghe'):
                self._add_zonghe_section(ws, row, data.get('zonghe', {}))

            # openpyxl 不支持真正的「自动列宽」，只能显式设定字符宽度：
            # A 列放标签（较窄），B 列放批注长文本（放宽），避免打开后满屏 ####
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 60

            wb.save(file_path)
            return True
        except Exception as e:
            # 导出属于用户主动触发的非关键路径，失败时不应让界面崩溃，
            # 统一吞掉异常并以返回值告知调用方
            print(f"Excel 导出失败: {e}")
            return False

    def _add_section(self, ws, start_row, title, data: dict) -> int:
        """添加键值区块，返回下一空行行号"""
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        for key, value in data.items():
            self._cell(ws, row, 1, str(key), align='left')
            self._cell(ws, row, 2, str(value), align='left')
            row += 1
        return row

    def _add_bazi_types_section(self, ws, start_row, bt: dict) -> int:
        """写入「命局类型」区块：日主强弱、格局、五行旺衰与用神喜忌。

        不能复用通用的 _add_section，因为 bazi_types 里的键需要中文改名、
        条件拼接（格局别名）和层级依赖（喜忌神依附用神），而非原样平铺。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            bt: bazi_types 子字典，含 strength / geju_type / geju_name /
                geju_desc / wuxing_summary / yongshen 等键

        Returns:
            int，下一个可写入的行号
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value='命局类型')
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        rows = []
        if bt.get('strength'):
            rows.append(('日主强弱', bt.get('strength')))
        if bt.get('geju_type'):
            # 格局别名（如「正官格」的通俗称呼）只在存在时才加括号后缀
            name = bt.get('geju_name', '')
            val = bt.get('geju_type') + (f"（{name}）" if name else '')
            rows.append(('格局类型', val))
            # 格局说明依附于格局类型，父项没有时不单独出现
            if bt.get('geju_desc'):
                rows.append(('格局说明', bt.get('geju_desc')))
        if bt.get('wuxing_summary'):
            rows.append(('五行旺衰', bt.get('wuxing_summary')))
        ys = bt.get('yongshen') or {}
        # 喜神/忌神是用神推导的衍生结果，用神未定则整组都不展示
        if ys.get('yongshen'):
            rows.append(('用神', f"{ys.get('yongshen')}（{ys.get('yongshen_name')}）"))
            if ys.get('xishen_names'):
                rows.append(('喜神', '、'.join(ys.get('xishen_names'))))
            if ys.get('jishen_names'):
                rows.append(('忌神', '、'.join(ys.get('jishen_names'))))

        for k, v in rows:
            self._cell(ws, row, 1, str(k), align='left')
            self._cell(ws, row, 2, str(v), align='left')
            row += 1
        if row == start_row + 1:  # 无内容
            # 游标没动说明一行都没写，仍要前进一行，否则下个区块标题会盖住本区块标题
            row += 1
        return row

    def _add_wuxing_section(self, ws, start_row, title, wuxing_data: dict) -> int:
        """写入「五行分析」区块。

        五行数据本身就是「五行名 -> 分数」的扁平字典，与通用键值区块结构一致，
        故直接委托 _add_section；保留独立方法是为了与其他章节调用形式统一，
        并预留后续单独定制样式的入口。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            title: 区块标题文案
            wuxing_data: wuxing 子字典，五行名到分数的映射

        Returns:
            int，下一个可写入的行号
        """
        return self._add_section(ws, start_row, title, wuxing_data)

    def _add_shishen_section(self, ws, start_row, title, shishen_data: dict) -> int:
        """写入「十神分析」区块。

        与五行同理，十神数据是「十神名 -> 分数」的扁平字典，直接委托 _add_section。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            title: 区块标题文案
            shishen_data: shishen 子字典，十神名到分数的映射

        Returns:
            int，下一个可写入的行号
        """
        return self._add_section(ws, start_row, title, shishen_data)

    def _add_yunshi_section(self, ws, start_row, data: dict) -> int:
        """写入「大运流年」区块：大运方向、起运说明、各步大运与逐年流年。

        这是一个虚拟章节，数据来自 dayun 与 liunian 两个并列的键，
        所以入参收的是整个 data 而非某个子字典。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            data: 完整排盘结果字典，读取 dayun.periods / dayun.direction /
                dayun.qiyun_text 与 liunian.years

        Returns:
            int，下一个可写入的行号；大运流年均为空时至少前进一行
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value='大运流年')
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        dayun = data.get('dayun', {}) or {}
        liunian = data.get('liunian', {}) or {}
        # 上游偶尔会传入非 dict 的旧结构，先做 isinstance 兜底再取子键
        periods = dayun.get('periods', []) if isinstance(dayun, dict) else []
        years = liunian.get('years', []) if isinstance(liunian, dict) else []

        if periods:
            self._cell(ws, row, 1, '大运', align='left')
            self._cell(ws, row, 2, str(dayun.get('direction', '')), align='left')
            row += 1
            qiyun = dayun.get('qiyun_text')
            if qiyun:
                self._cell(ws, row, 1, '起运', align='left')
                self._cell(ws, row, 2, str(qiyun), align='left')
                row += 1
            for p in periods:
                label = (f"第{p.get('period')}运 {p.get('ganzhi')} "
                         f"（{p.get('start_age')}-{p.get('end_age')}岁，"
                         f"{p.get('start_year')}-{p.get('end_year')}年）")
                self._cell(ws, row, 1, label, align='left')
                self._cell(ws, row, 2, str(p.get('analysis', '')), align='left')
                row += 1
        if years:
            # 流年小标题独占一行；B 列写空串是为了让该行也带上边框，与上下对齐
            self._cell(ws, row, 1, '流年', align='left')
            self._cell(ws, row, 2, '', align='left')
            row += 1
            for y in years:
                label = f"{y.get('year')}年 {y.get('ganzhi')}"
                self._cell(ws, row, 1, label, align='left')
                self._cell(ws, row, 2, str(y.get('analysis', '')), align='left')
                row += 1
        if row == start_row + 1:
            # 游标没动说明一行都没写，仍要前进一行，避免下个区块覆盖本区块标题
            row += 1
        return row

    def _add_yuncheng_section(self, ws, start_row, yc: dict) -> int:
        """写入「运程总结」区块：综述 + 事业/财运/健康/感情四维 + 标签。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            yc: yuncheng 子字典，含 overview / career / wealth / health /
                love / tags 等键，均为可选

        Returns:
            int，下一个可写入的行号；无任何内容时至少前进一行
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value='运程总结')
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        overview = yc.get('overview')
        if overview:
            self._cell(ws, row, 1, '综合', align='left')
            self._cell(ws, row, 2, str(overview), align='left')
            row += 1
        for key, label in (('career', '事业'), ('wealth', '财运'),
                           ('health', '健康'), ('love', '感情')):
            text = yc.get(key)
            if text:
                self._cell(ws, row, 1, label, align='left')
                self._cell(ws, row, 2, str(text), align='left')
                row += 1
        tags = yc.get('tags') or []
        if tags:
            self._cell(ws, row, 1, '标签', align='left')
            self._cell(ws, row, 2, '、'.join(str(t) for t in tags), align='left')
            row += 1
        if row == start_row + 1:
            # 游标没动说明一行都没写，仍要前进一行，避免下个区块覆盖本区块标题
            row += 1
        return row

    def _add_shensha_section(self, ws, start_row, title, shensha_list):
        """写入「神煞」区块：逐条列出神煞名称与释义。

        与其他区块不同，本方法不返回行号，调用方按 len(shensha_list)
        自行推算游标。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            title: 区块标题文案
            shensha_list: 神煞列表，元素为含 name / description 的字典

        Returns:
            None
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        for ss in shensha_list:
            self._cell(ws, row, 1, str(ss.get('name', '')), align='left')
            self._cell(ws, row, 2, str(ss.get('description', '')), align='left')
            row += 1

    def _add_analysis_section(self, ws, start_row, title, analysis_list):
        """写入「吉凶批注」区块：逐条批注，并按吉凶给类型单元格上底色。

        与其他区块不同，本方法不返回行号，调用方按 len(analysis_list)
        自行推算游标。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            title: 区块标题文案
            analysis_list: 批注列表，元素为含 type（吉/凶）与 text 的字典

        Returns:
            None
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        for an in analysis_list:
            # 用底色做吉凶的视觉编码：凶=朱红、吉=绿、其余（平/中性）=橙，
            # 这样在 Excel 里筛选浏览时无需逐字读类型
            type_color = 'FFC45545' if an.get('type') == '凶' else 'FF4CAF50' if an.get('type') == '吉' else 'FFFFA726'
            self._cell(ws, row, 1, str(an.get('type', '')), fill=type_color, align='center')
            self._cell(ws, row, 2, str(an.get('text', '')), align='left')
            row += 1

    def _add_ai_section(self, ws, start_row, ai: dict):
        """写入「龙虎山大师兄分析预测」区块：AI 生成的各维度解读。

        按 _AI_SECTIONS 固定顺序遍历，使区块次序不受 AI 返回字段顺序影响；
        每个维度先占一行小标题，其下每个条目单独成行并在 A 列打项目符号。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            ai: ai_analysis 子字典，键见 _AI_SECTIONS，值为字符串列表

        Returns:
            None
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value='龙虎山大师兄分析预测')
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        for key, title in _AI_SECTIONS:
            items = ai.get(key, []) or []
            # 防御：个别字段可能为字符串（如旧缓存数据），统一包成单元素列表，
            # 避免逐字符写入单元格。
            if isinstance(items, str):
                items = [items] if items.strip() else []
            elif not isinstance(items, (list, tuple)):
                items = [items] if items else []
            if not items:
                continue
            self._cell(ws, row, 1, title, align='left')
            self._cell(ws, row, 2, '', align='left')
            row += 1
            for it in items:
                self._cell(ws, row, 1, '•', align='center')
                self._cell(ws, row, 2, str(it), align='left')
                row += 1

    def _cell(self, ws, r, c, value, align='left', fill=None):
        """写入一个正文单元格并套用统一的字体、对齐与边框。

        所有区块的数据行都经由此方法落笔，是保证整张表样式一致的唯一出口。

        Args:
            ws: openpyxl 工作表对象
            r: 行号（从 1 开始）
            c: 列号（从 1 开始）
            value: 单元格内容，调用方应保证已转成字符串
            align: 对齐方式，取 self.styles 中的 'left' 或 'center'
            fill: 可选的 aRGB 背景色字符串；仅吉凶批注等需要临时着色的
                单元格才传，故不预建样式而是现场构造 PatternFill

        Returns:
            None
        """
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = self.styles['content']
        cell.alignment = self.styles[align]
        cell.border = self.styles['border']
        if fill:
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')

    def _add_zonghe_section(self, ws, start_row, z: dict):
        """写入「综合建议」区块：AI 融合八字/梅花/六壬三方结论后的统一判断。

        是整张表的收口区块，末尾附免责说明。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            z: zonghe 子字典，含 tri_method_overview / consistency_check /
                synthesis / unified_plan / key_timing（均为字符串列表）
                以及 disclaimer

        Returns:
            None
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value='综合建议（大师兄融合）')
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        # 顺序即「先摆三方结论 -> 再校验矛盾 -> 后给定论和方案」的推理链条，不可随意调换
        sections = [
            ('tri_method_overview', '三方概览'),
            ('consistency_check', '矛盾与印证'),
            ('synthesis', '综合定论'),
            ('unified_plan', '统一趋吉避凶方案'),
            ('key_timing', '关键时机与禁忌'),
        ]
        for key, title in sections:
            items = z.get(key) or []
            if not items:
                continue
            self._cell(ws, row, 1, title, align='left')
            self._cell(ws, row, 2, '', align='left')
            row += 1
            for it in items:
                self._cell(ws, row, 1, '•', align='center')
                self._cell(ws, row, 2, str(it), align='left')
                row += 1
        if z.get('disclaimer'):
            self._cell(ws, row, 1, '免责说明', align='left')
            self._cell(ws, row, 2, str(z.get('disclaimer')), align='left')
            row += 1

    def get_file_extension(self) -> str:
        """返回本导出器对应的文件扩展名。

        实现 BaseExporter.get_file_extension 契约，供上层拼接默认文件名
        和文件对话框过滤器使用。

        Returns:
            str，固定为 '.xlsx'
        """
        return '.xlsx'

    def _add_meihua_section(self, ws, start_row, mh: dict, mh_ai: dict):
        """写入「梅花易数卦象」区块：卦象要素 + AI 卦理解读。

        梅花的排盘数据与 AI 解读来自两个独立的键，用户可能只算了卦没跑 AI，
        故分成两个入参、各自判空后再决定是否渲染对应部分。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            mh: meihua_data 子字典，含 method / base_hex / changed_hex /
                hu_hex / ti_gong / yong_gong / hex_relation 等键
            mh_ai: meihua_ai 子字典，键见 _AI_SECTIONS，另有 final_verdict 结论

        Returns:
            None
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value='梅花易数卦象')
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        for key, label in (
            ('method', '起卦方法'), ('base_hex', '本卦'), ('changed_hex', '变卦'),
            ('hu_hex', '互卦'), ('ti_gong', '体卦'), ('yong_gong', '用卦'),
            ('ti_zhi', '体卦五行'), ('yong_zhi', '用卦五行'),
            ('hu_gong', '互卦五行'), ('bian_gong', '变卦五行'),
            ('hex_relation', '体用关系'),
        ):
            val = mh.get(key)
            if val:
                self._cell(ws, row, 1, label, align='left')
                self._cell(ws, row, 2, str(val), align='left')
                row += 1

        if mh_ai:
            # AI 解读前空出一行，并用与区块标题同款样式起一个二级小标题，
            # 把「卦象事实」与「AI 主观解读」在视觉上分开
            row += 1
            ws.merge_cells(f'A{row}:B{row}')
            sub = ws.cell(row=row, column=1, value='— 龙虎山大师兄梅花解读 —')
            sub.font = self.styles['header']
            sub.alignment = self.styles['center']
            sub.fill = self.styles['fill_gold']
            sub.border = self.styles['border']
            row += 1
            for key, title in _AI_SECTIONS:
                items = mh_ai.get(key, []) or []
                if not items:
                    continue
                self._cell(ws, row, 1, title, align='left')
                self._cell(ws, row, 2, '', align='left')
                row += 1
                for it in items:
                    self._cell(ws, row, 1, '•', align='center')
                    self._cell(ws, row, 2, str(it), align='left')
                    row += 1
            # final_verdict 是梅花特有的总断语，独立于 _AI_SECTIONS 各分项，放在最后压轴
            if mh_ai.get('final_verdict'):
                self._cell(ws, row, 1, '结论', align='left')
                self._cell(ws, row, 2, str(mh_ai.get('final_verdict')), align='left')
                row += 1

    def _add_liuren_section(self, ws, start_row, lr: dict, lr_ai: dict):
        """写入「大六壬起课」区块：课体要素 + AI 课理解读。

        结构与 _add_meihua_section 完全对称：排盘数据与 AI 解读来自两个
        独立的键，各自判空后再决定是否渲染对应部分。

        Args:
            ws: openpyxl 工作表对象
            start_row: 本区块标题所在行号
            lr: liuren_data 子字典，含 pan_date / si_ke / san_chuan / gate /
                yue_jiang / tian_jiang / shen_sha 等键
            lr_ai: liuren_ai 子字典，键见 _AI_SECTIONS，另有 final_verdict 结论

        Returns:
            None
        """
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value='大六壬起课')
        title_cell.font = self.styles['header']
        title_cell.alignment = self.styles['center']
        title_cell.fill = self.styles['fill_gold']
        title_cell.border = self.styles['border']

        row = start_row + 1
        for key, label in (
            ('pan_date', '公历日期'), ('si_ke', '四课'), ('san_chuan', '三传'),
            ('gate', '三传门法'), ('yue_jiang', '月将'),
            ('tian_jiang', '天将'), ('shen_sha', '神煞'),
        ):
            val = lr.get(key)
            if val:
                self._cell(ws, row, 1, label, align='left')
                self._cell(ws, row, 2, str(val), align='left')
                row += 1

        if lr_ai:
            # AI 解读前空出一行，并用与区块标题同款样式起一个二级小标题，
            # 把「课体事实」与「AI 主观解读」在视觉上分开
            row += 1
            ws.merge_cells(f'A{row}:B{row}')
            sub = ws.cell(row=row, column=1, value='— 龙虎山大师兄六壬解读 —')
            sub.font = self.styles['header']
            sub.alignment = self.styles['center']
            sub.fill = self.styles['fill_gold']
            sub.border = self.styles['border']
            row += 1
            for key, title in _AI_SECTIONS:
                items = lr_ai.get(key, []) or []
                if not items:
                    continue
                self._cell(ws, row, 1, title, align='left')
                self._cell(ws, row, 2, '', align='left')
                row += 1
                for it in items:
                    self._cell(ws, row, 1, '•', align='center')
                    self._cell(ws, row, 2, str(it), align='left')
                    row += 1
            # final_verdict 是六壬特有的总断语，独立于 _AI_SECTIONS 各分项，放在最后压轴
            if lr_ai.get('final_verdict'):
                self._cell(ws, row, 1, '结论', align='left')
                self._cell(ws, row, 2, str(lr_ai.get('final_verdict')), align='left')
                row += 1
